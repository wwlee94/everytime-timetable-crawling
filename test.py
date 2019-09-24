#-*- coding:utf-8 -*-
from selenium import webdriver
from bs4 import BeautifulSoup
from time import sleep
from openpyxl import Workbook,load_workbook

# Chrome의 경우 | 아까 받은 chromedriver의 위치를 지정해준다.
driver = webdriver.Chrome(executable_path=r'./chromedriver')

# url에 접근한다.
driver.get('https://everytime.kr/login')
# 암묵적으로 웹 자원 로드를 위해 3초까지 기다려 준다.
driver.implicitly_wait(5)

# 아이디/비밀번호를 입력해준다.
driver.find_element_by_name('userid').send_keys('*****')
driver.find_element_by_name('password').send_keys('*****')

# 로그인 버튼을 눌러주자.
driver.find_element_by_xpath('//*[@id="container"]/form/p[3]/input').click()

driver.get('https://everytime.kr/timetable')

#수업 목록에서 검색 클릭
driver.find_element_by_xpath('//*[@id="container"]/ul/li[1]').click()

#팝업창 닫기
sleep(2)
driver.find_element_by_xpath('//*[@id="sheet"]/ul/li[3]/a').click()

pre_count = 0
#스크롤 맨아래로 내리기
while True:
    #tr요소 접근
    element = driver.find_elements_by_css_selector("#subjects > div.list > table > tbody > tr")

    # tr 마지막 요소 접근
    result = element[-1]
    #마지막요소에 focus주기
    driver.execute_script('arguments[0].scrollIntoView(true);',result)
    sleep(2)

    #현재 접근한 요소의 갯수
    current_count = len(element)
    if pre_count == current_count:
        break
    #같지않다면
    pre_count = current_count


html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')

trs = soup.select('#subjects > div.list > table > tbody > tr')

results = []

for tr in trs:
    result=[]
    tds = tr.select('#subjects > div.list > table > tbody > tr > td')
    result.append(tds[0].text) #과목코드
    result.append(tds[1].text) #과목명
    result.append(tds[2].text) #교수
    result.append(tds[3].text) #강의시간
    result.append(tds[4].text) #강의실
    result.append(tds[5].text) #구분
    # result.append(tds[6].text) #학년
    result.append(tds[7].text) #학점
    result.append(tds[11].text) #수강대상
    results.append(result)
#값이 들어있다면!
if results:
    print("성공!!")
excel_column = 9
write_wb = Workbook()
write_ws = write_wb.create_sheet('result.xls')
for data in results:
    write_ws = write_wb.active
    write_ws.append(data)
write_wb.save('C:/Users/wwlee/Desktop/selenium/result.csv')



# Naver 페이 들어가기
# driver.get('https://order.pay.naver.com/home')
# html = driver.page_source
# soup = BeautifulSoup(html, 'html.parser')
# notices = soup.select('div.goods_item > div > a > p')
#
# for n in notices:
#     print(n.text.strip())
